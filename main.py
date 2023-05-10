
import sys
import os
import string
import re
import copy
import traceback
from decimal import Decimal
from shutil import copyfile

from PyQt5.QtWidgets import QApplication, QMessageBox, QFileDialog, QWidget, QMainWindow
from PyQt5.QtGui import QIcon


from ui.window import Ui_MainWindow
import resources.icon_rc
from worker import step1_worker
from worker import step2_worker
from error import contentError, exceptError, simpleError, baseError

from openpyxl import load_workbook


class extraWidgets(object):
    # 基础消息框
    def __init__(self, mainwindow):
        self.mainwindow = mainwindow
        self.infoIcon = QMessageBox.Icon(1)
        self.errorIcon = QMessageBox.Icon(3)
        self.questionIcon = QMessageBox.Icon(4)

    def errorMsg(self, emsg):
        baseBox = QMessageBox(self.mainwindow)
        baseBox.setIcon(self.errorIcon)
        baseBox.setText(str(emsg))
        baseBox.setDetailedText(emsg.detailedMSG)
        baseBox.exec()
  
    def fieleChoose(self, mode="select"):
        if mode == "select":
            dlog_fileChoose = QFileDialog(self.mainwindow)
            dlog_fileChoose.setFileMode(QFileDialog.ExistingFiles)
            dlog_fileChoose.setViewMode(QFileDialog.Detail)
            # 设置文件过滤器
            dlog_fileChoose.setNameFilter("Microsoft Excel files (*.xlsx)")
            if dlog_fileChoose.exec_():
                return dlog_fileChoose.selectedFiles()
        elif mode=="saveFile":
            dlog_fileChoose = QFileDialog.getSaveFileName(parent=self.mainwindow, caption="保存文件",
                filter="Microsoft Excel files (*.xlsx)")
            return dlog_fileChoose

class extraFunc:
    @staticmethod
    def guessSheetXY(sheet: any) -> dict:
        # 获取表格的title的最大探测行数
        maxRowDetection = 8
        maxColumnDetection = 8

        rowBegin = 1
        sheetXY = {}.fromkeys(("row", "column"), 0)
        while rowBegin <= maxRowDetection:
            columnBegin = 1
            for y in range(maxColumnDetection):
                if sheet.cell(row=rowBegin, column=columnBegin).value:
                    sheetXY["row"] = rowBegin
                    sheetXY["column"] = columnBegin
                    break
                columnBegin += 1
            if sheetXY["row"] and sheetXY["column"]:
                break
            rowBegin += 1
        return sheetXY

    @staticmethod
    def getRanged(desc: str)-> (range, baseError):
        if desc:
            if desc[0] in string.digits and '-' in desc:
                # 区间的规则
                try:
                    _r = desc.split("-")
                    assert len(_r) == 2, "规则长度错误"
                except Exception as e:
                    return None, exceptError(traceback.format_exc(), reason=str(e))
                else:
                    # 规则字符串中的w字符的处理
                    start_base = end_base = 1
                    start = _r[0]
                    end = _r[1]
                    if "w" in start or "W" in start:
                        start_base = 10000
                        start = start.replace("w", "").replace("W", "")
                    if "w" in end or "W" in end:
                        end_base = 10000
                        end = end.replace("w", "").replace("W", "")

                    # 规则区间数字大小限制
                    try:
                        assert int(float(start) *  start_base) < int(float(end) * end_base), "规则错误（a-b;a<b）"
                    except Exception as e:
                        return None, exceptError(traceback.format_exc(), reason=str(e))
                    
                    try:
                        assert (float(start) * start_base).is_integer() is True, "区间开始不能有小数"
                        assert (float(end) * end_base).is_integer() is True, "区间结束不能有小数"
                    except Exception as e:
                        return None, exceptError(traceback.format_exc(), reason=str(e))

                    return range(int(float(start) * start_base),  int(float(end) * end_base)), None
            else:
                # 金额base
                base = 1
                p = desc
                if  "w" in desc or "W" in desc:
                    base = 10000
                    p = p.replace("w", "").replace("W", "")
                # >,>=,<,<=,= 的情况
                matched = False
                if ">=" in desc and not matched:
                    matched = True
                    # 大于等于情况
                    p = p.replace(">=", "")
                    rge = range(int(float(p)* base) - 1, 100000000)

                if ">" in desc and not matched:
                    matched = True
                    # 大于情况
                    p = p.replace(">", "")
                    rge = range(int(float(p)* base) + 1, 100000000)

                if "<=" in desc and not matched:
                    matched = True
                    # 小于等于情况
                    p = p.replace("<=", "")
                    rge = range(0, int(float(p)* base) + 1)

                if "<" in desc and not matched:
                    matched = True
                    # 小于情况
                    p = p.replace("<", "")
                    rge = range(0, int(float(p)* base))

                if "=" in desc and not matched:
                    matched = True
                    # 等于的情况
                    p = p.replace("=", "")
                    rge =  range(int(float(p)* base), int(float(p)* base)+1)
                try:
                    assert (float(p) * base).is_integer() is True, "区间不能有小数"
                except Exception as e:
                    return None, exceptError(traceback.format_exc(), reason=str(e))
                return rge, None
        else:
            return None, simpleError("规则为空")

        return None, simpleError("规则格式未知")

class excelSourceProcessWorker(extraFunc):
    def setupWorker1(self, s):

        # 配置文件是否分析完成
        self.configFileAnalysisDone = False
        self.sourceFileAnalysisDone = False
        self.sourceFileAnalysisResult = ""

        self.step1ConfigFileChooseBT.clicked.connect(lambda: self.configFileHandler())
        self.step1SourceFileChooseBT.clicked.connect(lambda: self.sourceFileHandler())

        # 进度条值范围
        self.progressBar.setRange(0, 100)

        # 分析
        self.step1ProcessBT.clicked.connect(lambda: self.step1ProcessWorker())
        self.processWorkerThread = step1_worker.step1ProcessWorker()
        self.processWorkerThread.signalProgress.connect(self.handleProgress)
        self.processWorkerThread.signalError.connect(self.handleError)

        # 导出
        self.step1ExportBT.clicked.connect(lambda: self.step1ExportWorker())

    def analysisConfigFile(self):
        try:
            excel_file = self.step1ConfigFileChooseRV.text()
            if not os.path.exists(excel_file):
                self.extr_widget.errorMsg(simpleError("配件、折扣信息配置Excel文件：【{}】不存在".format(excel_file)))
                return

            configs = {}.fromkeys(["peijian", "discount", "activity"], {})

            try:
                conf_excel = load_workbook(filename=excel_file, read_only=True)
            except Exception as e:
                except_error = exceptError(
                    traceback.format_exc(),
                    reason="配件、折扣信息配置Excel文件：【{}】打开失败".format(excel_file)
                )
                self.extr_widget.errorMsg(except_error)
                return
            else:
                self.comboBox_3.clear()
                self.comboBox.clear()
                self.comboBox_2.clear()
                # todo 默认按照预定的sheetname进行数据查找
                index = 0
                for sheet_name in conf_excel.sheetnames:
                    self.comboBox_3.insertItem(index, sheet_name)
                    self.comboBox_3.setItemData(index, sheet_name)
                    self.comboBox.insertItem(index, sheet_name)
                    self.comboBox.setItemData(index, sheet_name)
                    self.comboBox_2.insertItem(index, sheet_name)
                    self.comboBox_2.setItemData(index, sheet_name)
                    index += 1
                # ======================================================================================================
                # 产品及配件价格表
                if "产品及配件价格表" not in conf_excel.sheetnames:
                    self.extr_widget.errorMsg(simpleError("【产品及配件价格表】不存在，请检查"))
                    return
                else:
                    sheet_price = conf_excel["产品及配件价格表"]
                    # 样例数据（配件价格表）
                    # {
                    #     "餐具": {"GuiGe": ["5套"], "JiaGe": 5},
                    #     "数字蜡烛": {"GuiGe": ["数字0", "数字1"], "JiaGe": 5},
                    # }
                    valid_HPMC = ["餐具", "派对生日蜡烛", "生日蜡烛", "生日帽", "数字蜡烛", "运费/差价"]
                    peijian_config = {}.fromkeys(valid_HPMC)
                    guessXY = self.guessSheetXY(sheet_price)

                    if guessXY["row"] and guessXY["column"]:
                        index = 1
                        while True:
                            HuoPinMingCheng = sheet_price.cell(row= guessXY["row"] + index, column=guessXY["column"] + 1).value
                            GuiGe = sheet_price.cell(row= guessXY["row"] + index , column=guessXY["column"] + 2).value
                            JiaGe = sheet_price.cell(row= guessXY["row"] + index, column=guessXY["column"] + 3).value
                            if HuoPinMingCheng and GuiGe and JiaGe:
                                if HuoPinMingCheng in peijian_config:
                                    if peijian_config[HuoPinMingCheng] is None:
                                        peijian_config[HuoPinMingCheng] = {
                                            "GuiGe": [], "JiaGe": Decimal(0).quantize(Decimal("0.00"))
                                        }
                                        peijian_config[HuoPinMingCheng]["GuiGe"].append(GuiGe)
                                        peijian_config[HuoPinMingCheng]["JiaGe"]=Decimal(JiaGe).quantize(Decimal("0.00"))
                                    else:
                                        peijian_config[HuoPinMingCheng]["GuiGe"].append(GuiGe)
                            else:
                                break
                            index += 1
                    else:
                        self.extr_widget.errorMsg(simpleError("表格未找到任何值，可能是个空表格？"))
                        return
                    configs["peijian"] = peijian_config
                # ======================================================================================================
                
                # ------------------------------------------佳满勇气--各渠道折扣--------------------------------------------
                # 佳满勇气--各渠道折扣
                # 折扣信息描述语言使用规则语言这一列
                # 样例数据（配件价格表）
                # {
                #     "福多多": [(<RangeObject>, "%90"), (<RangeObject>, "%78")],
                #     "蛋糕叔叔": [(<RangeObject>, "%78")]
                # }
                discount_config = {}
                if "佳满勇气--各渠道折扣" not in conf_excel.sheetnames:
                    self.extr_widget.errorMsg(simpleError("【佳满勇气--各渠道折扣】不存在，请检查"))
                    return
                else:
                    sheet_discount = conf_excel["佳满勇气--各渠道折扣"]
                    guessXY = self.guessSheetXY(sheet_discount)
                    if guessXY["row"] and guessXY["column"]:
                        index1 = 1
                        while True:
                            QuDaoMingCheng = sheet_discount.cell(row=guessXY["row"] + index1, column=guessXY["column"]).value
                            GuiZeYuYan = sheet_discount.cell(row=guessXY["row"] + index1, column=guessXY["column"]+2).value
                            ZheKou = sheet_discount.cell(row=guessXY["row"] + index1, column=guessXY["column"]+3).value
                            
                            # 折扣没有填写，则默认没有折扣
                            if ZheKou is None:
                                ZheKou = 1.0

                            if QuDaoMingCheng:
                                range_obj, error_msg = self.getRanged(GuiZeYuYan)
                                if error_msg is not None:
                                    self.extr_widget.errorMsg(error_msg)
                                else:
                                    # 向数组中添加range对象和折扣信息
                                    if QuDaoMingCheng in discount_config:
                                        discount_config[QuDaoMingCheng].append((range_obj,ZheKou))
                                    else:
                                        discount_config[QuDaoMingCheng] = []
                                        discount_config[QuDaoMingCheng].append((range_obj,ZheKou))
                            else:
                                break
                            index1 += 1
                    else:
                        self.extr_widget.errorMsg(simpleError("表格未找到任何值，可能是个空表格？"))
                        return

                    configs["discount"] = discount_config

                # ------------------------------------------佳满勇气--各渠道折扣--------------------------------------------


                # ¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥各渠道活动¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥
                # 各渠道活动
                # ¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥各渠道活动¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥¥
                conf_excel.close()

            self.configFileAnalysisDone = True
            return configs
        except Exception as e:
            self.extr_widget.errorMsg(exceptError(traceback.format_exc(), reason="分析配置Excel发生了未知错误"))
            return
    
    def step1ProcessWorker(self):
        if self.configFileAnalysisDone:
            if self.step1SourceDataSheetSelect.currentData() is None:
                self.extr_widget.errorMsg(simpleError("请选择：源Excel表格对应的月份数据"))
                return
            params = {
                "source_file": self.step1SourceFileChooseRV.text(),
                "source_sheet": self.step1SourceDataSheetSelect.currentData(),
                "guessSheetXY": self.guessSheetXY,
                "configs": self.analysisConfigFile()
            }
            self.processWorkerThread.runParams(**params)
            self.processWorkerThread.start()
            self.progressBar.reset()

        else:
            self.extr_widget.errorMsg(simpleError("配件信息配置分析出错，请检查后在继续"))
            return

    def step1ExportWorker(self):
        if self.sourceFileAnalysisDone:
            filename = self.extr_widget.fieleChoose(mode="saveFile")
            if filename[0]:
                if os.path.exists(self.sourceFileAnalysisResult):
                    copyfile(self.sourceFileAnalysisResult, filename[0])
                else:
                    self.extr_widget.errorMsg(simpleError("分析结果文件不存在！！"))
                    return
        else:
            self.extr_widget.errorMsg(simpleError("源数据文件表格还未分析完成"))
            return 

    def configFileHandler(self):
        filename = self.extr_widget.fieleChoose()
        if filename:
            self.step1ConfigFileChooseRV.setText(filename[0])
            self.analysisConfigFile()

    def sourceFileHandler(self):
        filename = self.extr_widget.fieleChoose()
        
        if filename:
            self.step1SourceFileChooseRV.setText(filename[0])
        
            if not os.path.exists(self.step1SourceFileChooseRV.text()):
                self.extr_widget.errorMsg(simpleError("Excel文件：【{}】不存在".format(self.step1SourceFileChooseRV.text())))
                return

            sfile = load_workbook(filename=self.step1SourceFileChooseRV.text(), read_only=True)
            self.step1SourceDataSheetSelect.clear()
            self.step1SourceDataSheetSelect.insertItem(0, "选择月份数据")
            self.step1SourceDataSheetSelect.setItemData(0, None)
            index = 1
            for sname in sfile.sheetnames:
                self.step1SourceDataSheetSelect.insertItem(index, sname)
                self.step1SourceDataSheetSelect.setItemData(index, sname)
                index += 1
            self.step1SourceDataSheetSelect.setCurrentIndex(0)
            sfile.close()

class excelAnalysisWroker(extraFunc):
    def setupWorker2(self, ss):
        
        self.third_configs = []
        self.source_configs = {}
        
        self.thirdAnalysisDone = False
        self.thirdAnalysisResult = ""

        # 进度条值范围
        self.analysisProgressBar.setRange(0, 100)

        self.sourceFileChooseBT.clicked.connect(lambda: self.sourceFileChooseHandler())
        # signal `currentTextChanged`
        self.sourceWS.currentTextChanged.connect(lambda text: self.workSheetChangeHandler(text, filesource="source"))

        self.thirdFileChooseBT.clicked.connect(lambda: self.thirdFileChooseHandler())
        self.thirdWS.currentTextChanged.connect(lambda text: self.workSheetChangeHandler(text, filesource="third"))

        self.thirdInfoAddBT.clicked.connect(lambda: self.addThirdInfo())
        self.thirdInfoEmptyBT.clicked.connect(lambda: self.emptyThirdInfo())

        self.beginAnalysisBT.clicked.connect(lambda: self.beginAnalysis())
        self.analysisWorkerThread = step2_worker.step2ProcessWorker()
        self.analysisWorkerThread.signalProgress.connect(self.handleProgress)
        self.analysisWorkerThread.signalError.connect(self.handleError)
        self.analysisWorkerThread.signalConsole.connect(self.outPutMsg)
        self.exportResultBT.clicked.connect(lambda: self.exportAnalysisResult())

        self.exportResultBT.clicked.connect(lambda: self.analysisResultExport())

    def analysisResultExport(self):
        if self.thirdAnalysisDone:
            filename = self.extr_widget.fieleChoose(mode="saveFile")
            if filename[0]:
                if os.path.exists(self.thirdAnalysisResult):
                    copyfile(self.thirdAnalysisResult, filename[0])
                else:
                    self.extr_widget.errorMsg(simpleError("分析结果文件不存在！！"))
                    return
        else:
            self.extr_widget.errorMsg(simpleError("源数据文件表格还未分析完成"))
            return 

    def beginAnalysis(self):
        work_sheet = self.sourceWS.currentData()
        fname = self.sourceFileChooseReView.text()

        if fname == "【选择源文件】":
            self.extr_widget.errorMsg(simpleError("请选择公司配置所在的文件"))
            return

        if work_sheet == "选择WorkSheet":
            self.extr_widget.errorMsg(simpleError("请选择worksheet"))
            return
        key_info = self._sourceFileValidCheck(fname, work_sheet)
        if not key_info["amount_column"] or not key_info["order_column"]:
            return

        self.source_configs = {
            "file_name": fname,
            "work_sheet": work_sheet,
        }
        self.source_configs.update(key_info)

        # 清空信息
        self.analysisErrorMsg.setText("")
        params = {
            "source_configs": self.source_configs,
            "third_configs": self.third_configs,
            "guessSheetXY": self.guessSheetXY
        }
        self.analysisWorkerThread.runParams(**params)
        self.analysisWorkerThread.start()
        self.analysisProgressBar.reset()
    
    def exportAnalysisResult(self):
        return

    def sourceFileChooseHandler(self):
        filename = self.extr_widget.fieleChoose()

        if filename:
            self.sourceFileChooseReView.setText(filename[0])

        try:
            swb = load_workbook(filename=filename[0], read_only=True)
        except:
            pass
        else:
            self.sourceWS.clear()
            self.sourceWS.insertItem(0, "选择WorkSheet")
            self.sourceWS.setItemData(0, None)
            index = 1
            for sn in swb.sheetnames:
                self.sourceWS.insertItem(index, sn)
                self.sourceWS.setItemData(index, sn)
                index += 1 
            self.sourceWS.setCurrentIndex(0)
            swb.close()

    def _sourceFileValidCheck(self, file_name, sheet_name):
        key_info = {}.fromkeys(["order_column", "order_name", "amount_name", "amount_column"])
        try:
            swb = load_workbook(filename=file_name, read_only=True)
        except:
            self.extr_widget.errorMsg(simpleError("文件打开失败（%s）".format(file_name)))
            return
        else:
            if not sheet_name:
                self.extr_widget.errorMsg(simpleError("请选择公司数据源所在的sheet名称"))
                return key_info
            work_sheet = swb[sheet_name]

            guessXY = self.guessSheetXY(work_sheet)
            if guessXY["row"] and guessXY["column"]:
                column_start = guessXY["column"]
                while True:
                    title_name = work_sheet.cell(row=guessXY["row"], column=column_start).value
                    if title_name == "网店订单号":
                        key_info["order_column"] = column_start
                        key_info["order_name"] = "网店订单号"
                    if title_name == "结算金额":
                        key_info["amount_column"] = column_start
                        key_info["amount_name"] = "结算金额"
                    if not title_name:  break
                    column_start += 1
            else:
                self.extr_widget.errorMsg(simpleError("表格内容为空？"))
                return key_info

        return key_info

    def workSheetChangeHandler(self, sheet_name, filesource=""):
        if sheet_name:
            if sheet_name == "选择WorkSheet":
                return
            else:
                if filesource == 'source':
                    fname = self.sourceFileChooseReView.text()
                    # 检查表格中的订单ID以及订单总价字段
                    result_info = self._sourceFileValidCheck(fname, sheet_name)
                    if not result_info["order_column"] or not result_info["amount_column"]:
                        self.extr_widget.errorMsg(simpleError("表格内容不包含网店订单号或者结算金额"))
                        return

                elif filesource == "third":
                    fname = self.thirdFileChooseReView.text()
                    try:
                        twb = load_workbook(filename=fname, read_only=True)
                    except:
                        pass
                    third_ws = twb[sheet_name]
                    guessXY = self.guessSheetXY(third_ws)
                    if guessXY["row"] and guessXY["column"]:
                        self.thirdOrderIDColumn.clear()
                        self.thirdOrderIDColumn.insertItem(0, "请选择列信息")
                        self.thirdOrderIDColumn.setItemData(0, "请选择列信息")
                        
                        self.thirdOrderAmountColumn.clear()
                        self.thirdOrderAmountColumn.insertItem(0, "请选择列信息")
                        self.thirdOrderAmountColumn.setItemData(0, "请选择列信息")

                        start_column = guessXY["column"]
                        index = 1
                        while True:
                            title_name = third_ws.cell(row=guessXY["row"], column=start_column).value
                            if not title_name: break
                            self.thirdOrderIDColumn.insertItem(index, title_name)
                            self.thirdOrderIDColumn.setItemData(index, start_column)
                            self.thirdOrderAmountColumn.insertItem(index, title_name)
                            self.thirdOrderAmountColumn.setItemData(index, start_column)
                            start_column += 1
                            index += 1
                    else:
                        self.extr_widget.errorMsg(simpleError("表格为空？"))
                        return
        else:
            return

    def thirdFileChooseHandler(self):
        filename = self.extr_widget.fieleChoose()

        if filename:
            self.thirdFileChooseReView.setText(filename[0])

        try:
            twb = load_workbook(filename=filename[0], read_only=True)
        except:
            pass
        else:
            self.thirdWS.clear()
            self.thirdWS.insertItem(0, "选择WorkSheet")
            self.thirdWS.setItemData(0, None)
            index = 1
            for sn in twb.sheetnames:
                self.thirdWS.insertItem(index, sn)
                self.thirdWS.setItemData(index, sn)
                index += 1 
            self.thirdWS.setCurrentIndex(0)

            twb.close()

    def addThirdInfo(self):
        if not self.thirdOrderIDColumn.currentData() or not self.thirdOrderAmountColumn.currentData():
            self.extr_widget.errorMsg(simpleError("请选择好配置信息"))
            return
        if not self.thirdName.text():
            self.extr_widget.errorMsg(simpleError("请输入渠道名称"))
            return
        template_config = {}.fromkeys(["file_name", "work_sheet", "order_column", "amount_column"])
        template_config["file_name"]      = self.thirdFileChooseReView.text()
        template_config["work_sheet"]    = self.thirdWS.currentData()
        template_config["order_column"]  = self.thirdOrderIDColumn.currentData()
        template_config["order_name"]  = self.thirdOrderIDColumn.currentText()
        template_config["amount_column"] = self.thirdOrderAmountColumn.currentData()
        template_config["amount_name"] = self.thirdOrderAmountColumn.currentText()
        template_config["third_name"]    = self.thirdName.text()

        self.third_configs.append(template_config)
        html_fixed = "<html><body><p>要分析的渠道信息：</p><ul>{content}</ul></body></html>"
        ul_content = ""
        for cfg in self.third_configs:
            item = "<li>【{}】【{}】【{}】【{}】【{}】</li>".format(
                cfg["third_name"], os.path.basename(cfg["file_name"]), cfg["work_sheet"], cfg["order_name"], cfg["amount_name"])
            ul_content += item

        self.thirdInfoListReview.setHtml(html_fixed.format(content=ul_content))

    def emptyThirdInfo(self):
        self.third_configs = []
        self.thirdInfoListReview.setHtml("<html></body>")

class toolWindow(QMainWindow, Ui_MainWindow, excelSourceProcessWorker, excelAnalysisWroker):
    def __init__(self, parent=None):
        super(toolWindow, self).__init__()

        # 初始化window对象
        self.setupUi(self)
        self.setupWorker1(self)
        self.setupWorker2(self)

        self.extr_widget = extraWidgets(self)
    
    # 以下是公共方法
    def handleError(self, error):
        self.extr_widget.errorMsg(error)
        return

    def outPutMsg(self, out_msg):
        self.analysisErrorMsg.append(out_msg)

    def handleProgress(self, worker, progres):
        if worker == "step1ProcessWorker":
            self.processTitle.setText(progres[0])
            self.progressBar.setValue(progres[1])

            if progres[1] == 100:
                self.sourceFileAnalysisDone = True
                self.sourceFileAnalysisResult = progres[2]

        if worker == "analysisWorker":
            self.analysisProcessTitle.setText(progres[0])
            self.analysisProgressBar.setValue(progres[1])
            if progres[1] == 100:
                self.thirdAnalysisDone = True
                self.thirdAnalysisResult = progres[2]


if __name__ == "__main__":
    app = QApplication([])
    app.setWindowIcon(QIcon(":/image/icon/main.png"))
    w = toolWindow()
    w.show()
    sys.exit(app.exec_())