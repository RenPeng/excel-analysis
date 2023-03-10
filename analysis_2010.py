
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename='哈哈哈.xlsx')
worksheets = workbook.sheetnames
self_ = workbook[worksheets[0]]
other = workbook[worksheets[1]]

s = {}
o = {}
right_s = {}
right_o = {}
diff_s = {}
diff_o = {}
error_s = {}
error_o = {}

for i in range(self_.max_row):
    row_ = self_.row(i)
    key = str(row_[0].value).strip()
    value = row_[1].value
    if key:
        if key in s:
            s[key].append(float(value))
        else:
            s[key] = []
            s[key].append(float(value))

for i in range(other.nrows):
    row_ = other.row(i)
    key = str(row_[0].value).strip()
    value = row_[1].value
    if key:
        if key in o:
            o[key].append(float(value))
        else:
            o[key] = []
            o[key].append(float(value))
 
for k,v in s.items():
    if k not in o:
        if k not in diff_s:
            diff_s[k] = []
            diff_s[k].extend(v)
        else:
            diff_s[k].extend(v)
        print(f"订单:{k}, 总价:{sum(v)}, 不再他方数据表中")
    else:
        if sum(v) != sum(o[k]):
            print(f"订单:{k}, 不一致, 我方总价{sum(v)}, 他方{sum(o[k])}")
            if k not in error_s:
                error_s[k] = []
                error_s[k].extend(v)
            else:
                error_s[k].extend(v)
        else:
            if k not in right_s:
                right_s[k] = []
                right_s[k].extend(v)
            else:
                right_s[k].extend(v)

print("-_"*20)
for k,v in o.items():
    if k not in s:
        if k not in diff_o:
            diff_o[k] = []
            diff_o[k].extend(v)
        else:
            diff_o[k].extend(v)
        print(f"订单:{k}, 总价:{sum(v)}, 不再我方数据表中")
    else:

        if sum(v) != sum(s[k]):
            print(f"订单:{k}, 不一致, 他方总价{sum(v)}, 我方{sum(s[k])}")
            if k not in error_o:
                error_o[k] = []
                error_o[k].extend(v)
            else:
                error_o[k].extend(v)
        else:
            if k not in right_o:
                right_o[k] = []
                right_o[k].extend(v)
            else:
                right_o[k].extend(v)

wb = xlwt.Workbook()
ws = wb.add_sheet('嘻嘻嘻')

style_error = xlwt.easyxf('pattern: pattern solid, fore_colour  yellow')
style_diff = xlwt.easyxf('pattern: pattern solid, fore_colour  orange')

sort_right_s = dict(sorted(right_s.items(), key=lambda item: item[0]))
sort_right_o = dict(sorted(right_o.items(), key=lambda item: item[0]))
sort_diff_s = dict(sorted(diff_s.items(), key=lambda item: item[0]))
sort_diff_o = dict(sorted(diff_o.items(), key=lambda item: item[0]))
sort_error_s = dict(sorted(error_s.items(), key=lambda item: item[0]))
sort_error_o = dict(sorted(error_o.items(), key=lambda item: item[0]))


if len(error_o) != len(error_s):
    print('不一致的订单金额两边表格有误，请检查')
    exit()

row_start = 0
row_end_1 = 0
row_end_2 = 0



for k,v in sort_error_s.items():
    for v_ in v:
        ws.write(row_end_1, 0, k,  style_error)
        ws.write(row_end_1, 1, v_, style_error)
        row_end_1 += 1

    if k in sort_error_o:
        for v__ in sort_error_o[k]:
            ws.write(row_end_2, 2, k, style_error)
            ws.write(row_end_2, 3, v__, style_error)
            row_end_2 += 1

    row_end_1 = row_end_2 = max(row_end_1, row_end_2)

for k,v in sort_diff_s.items():
    for v_ in v:
        ws.write(row_end_1, 0, k, style_diff)
        ws.write(row_end_1, 1, v_, style_diff)
        row_end_1 += 1

    if k in sort_diff_o:
        for v__ in sort_diff_o[k]:
            ws.write(row_end_2, 2, k, style_diff)
            ws.write(row_end_2, 3, v__,style_diff)
            row_end_2 += 1

    row_end_1 = row_end_2 = max(row_end_1, row_end_2)

for k,v in sort_right_s.items():
    for v_ in v:
        ws.write(row_end_1, 0, k)
        ws.write(row_end_1, 1, v_)
        row_end_1 += 1

    if k in sort_right_o:
        for v__ in sort_right_o[k]:
            ws.write(row_end_2, 2, k)
            ws.write(row_end_2, 3, v__)
            row_end_2 += 1

    row_end_1 = row_end_2 = max(row_end_1, row_end_2)

wb.save('哈哈哈-1.xlsx')
print('\n\n')
input("文件生成完毕，按任意键退出")
