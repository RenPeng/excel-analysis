import traceback
import datetime
from decimal import Decimal
import re
import os

from openpyxl import load_workbook
from PyQt5.QtCore import QThread, pyqtSignal

from error import simpleError, contentError, exceptError, baseError

class step1ProcessWorker(QThread):
    # str: workername, list: [0: 阶段, 1:进度, 2: other]
    # progres = ["stage-desc", "stage", "other"]
    signalProgress = pyqtSignal(str, list)
    # 0: baseError
    signalError = pyqtSignal(baseError)

    def runParams(self, **params):
        self.params = params
        # params = {
        #     "source_file": self.step1SourceFileChooseRV.text(),
        #     "source_sheet": self.step1SourceDataSheetSelect.currentData(),
        #     "guessSheetXY": self.guessSheetXY,
        #     "configs": configs
        # }

    def run(self):
        try:
            try:
                sfile = load_workbook(filename=self.params["source_file"])
                sheet = sfile[self.params["source_sheet"]]
                self.signalProgress.emit("step1ProcessWorker",
                    ["【配置处理】【{}】打开Excel文件".format(self.params["source_sheet"]), 1])
            except:
                self.signalError.emit(exceptError(traceback.format_exc(), reason="源文件打开失败"))
                return

            # 计算表格的起始位置
            start_point = self.params["guessSheetXY"](sheet)
            # 折扣、配件价格配置信息
            configs = self.params["configs"]

            # 首先插入要添加的列
            for title in reversed(["货款合计Fixed", "配件费用", "折扣", "结算金额"]):
                sheet.insert_cols(11)
                sheet.cell(row=start_point["row"], column=11, value=title)

            # 进度信息推送
            self.signalProgress.emit("step1ProcessWorker",
                ["【配置处理】【新增数据列】货款合计Fixed、配件费用、折扣 、结算金额", 2])

            # 确定需要被处理信息坐在的列
            OrderStatus = OrderComment = 0
            OrderCreateTime = OrderDeliveryTime = 0
            OrderDetail = OrderItemTotal = OrderItemTotalFixed = OrderTotal = OrderExtra =  0
            QuDao = QuDaoDiscount =  0
            column_start = start_point["column"]
            while True:
                title_name = sheet.cell(row=start_point["row"], column=column_start).value
                if title_name =="订单状态": OrderStatus = column_start
                if title_name =="追加备注": OrderComment = column_start

                if title_name =="下单时间": OrderCreateTime = column_start
                if title_name =="承诺发货时间": OrderDeliveryTime = column_start

                if title_name =="货品摘要": OrderDetail = column_start
                if title_name =="货款合计": OrderItemTotal = column_start
                if title_name =="货款合计Fixed": OrderItemTotalFixed = column_start
                if title_name =="结算金额": OrderTotal = column_start
                if title_name =="配件费用": OrderExtra = column_start
                if title_name =="应收邮资": OrderPostage = column_start

                if title_name =="销售渠道": QuDao = column_start
                if title_name =="折扣": QuDaoDiscount = column_start
                column_start += 1
                if not title_name:
                    break

                # 进度信息推送
                self.signalProgress.emit("step1ProcessWorker",
                    ["【配置处理】【列信息获取】计算{} 所在的列".format(title_name), 3])

            rowBegin = start_point["row"]

            # 配件价格
            peijian_restr = "|".join(configs["peijian"].keys())
            # 渠道的销售总额记录
            QuDaoTotal = {}
            delete_rows = []
            for row_index in range(sheet.max_row):
                if row_index == 0:
                    rowBegin += 1
                    continue

                current_percent = int((row_index/(sheet.max_row * 2)) * 100)

                self.signalProgress.emit("step1ProcessWorker",[
                    "【数据处理】【第{}行】分析【订单状态】中取消的订单，【追加备注】中带有（取消/二次配送/不接算）的订单".format(rowBegin), 
                    current_percent]
                )

                # 需求1：删除【订单状态】中取消的订单，【追加备注】中带有（取消/二次配送/不接算）的订单
                osv_ = sheet.cell(row=rowBegin, column=OrderStatus).value
                ocv_ = sheet.cell(row=rowBegin, column=OrderComment).value
                if osv_ :
                    if "取消" in osv_:
                        delete_rows.append(rowBegin)
                if ocv_:
                    if ("取消" in ocv_ or "不结算" in  ocv_ or "二次配送" in ocv_):
                        delete_rows.append(rowBegin)

                # 需求2：【下单时间】【承诺发货时间】拆分出日期，删掉时间部分
                octv__ = sheet.cell(row=rowBegin, column=OrderCreateTime).value
                odtv__ = sheet.cell(row=rowBegin, column=OrderDeliveryTime).value

                if octv__:
                    if type(octv__) == datetime.datetime:
                        dt1 = octv__
                    elif type(octv__) == str:
                        dt1 = datetime.datetime.strptime(octv__, "%Y-%m-%d %H:%M:%S")
                    sheet.cell(row=rowBegin, column=OrderCreateTime, value=dt1.strftime("%Y/%m/%d"))

                # 推送进度数据
                self.signalProgress.emit("step1ProcessWorker", [
                    "【数据处理】【第{}行】拆分出日期，删掉时间部分".format(rowBegin), current_percent]
                )

                if odtv__:
                    if type(octv__) == datetime.datetime:
                        dt2 = octv__
                    elif type(octv__) == str:
                        dt2 = datetime.datetime.strptime(odtv__, "%Y-%m-%d %H:%M:%S")
                    sheet.cell(row=rowBegin, column=OrderDeliveryTime, value=dt2.strftime("%Y/%m/%d"))

                # 需求3：拆分货品摘要，计算货款合计、配件费用
                odv___ = sheet.cell(row=rowBegin, column=OrderDetail).value
                oitv___ = sheet.cell(row=rowBegin, column=OrderItemTotal).value
                
                # 推送进度数据
                self.signalProgress.emit("step1ProcessWorker", [
                    "【数据处理】【第{}行】拆分货品摘要，计算货款合计、配件费用".format(rowBegin), current_percent]
                )
                # 初始化值（货款合计Fixed）
                if oitv___:
                    sheet.cell(row=rowBegin, column=OrderItemTotalFixed, value=Decimal(oitv___).quantize(Decimal("0.00")))

                if odv___:
                    for item in odv___.split(","):
                        if re.findall(peijian_restr, item):
                            # 匹配货品和数量
                            mapped = re.search("(.*)\[.*\]\((.*)\)", item)
                            if mapped:
                                mapped_group = mapped.groups()
                                try:
                                    price = configs["peijian"][mapped_group[0]]["JiaGe"]
                                    number = int(mapped_group[1])
                                    # 配件费用修改
                                    # 现有的值
                                    oev___ = sheet.cell(row=rowBegin, column=OrderExtra).value
                                    if not oev___:
                                        oev_cur = Decimal(0).quantize(Decimal("0.00"))
                                    else:
                                        oev_cur = Decimal(oev___).quantize(Decimal("0.00"))
                                    sheet.cell(row=rowBegin, column=OrderExtra, value=oev_cur + price * number)

                                    # 修改货款合计（保存到货款合计Fixed）
                                    oitfv_cur = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value
                                    sheet.cell(row=rowBegin, column=OrderItemTotalFixed, value=oitfv_cur - price * number)
                                except Exception as e:
                                    self.signalError.emit(exceptError(traceback.format_exc(), reason="修改货款合计（保存到货款合计Fixed）出错"))
                            else:
                                self.signalError.emit(contentError("【货品摘要】配件信息格式有误", ""))

                # 需求4：计算渠道折扣
                # QuDaoDiscount
                qdv____ = sheet.cell(row=rowBegin, column=QuDao).value
                oitfv____ = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value

                if qdv____  not in QuDaoTotal:
                    QuDaoTotal[qdv____] = Decimal(oitfv____).quantize(Decimal("0.00"))
                else:
                    if oitfv____:
                        QuDaoTotal[qdv____] += Decimal(oitfv____).quantize(Decimal("0.00"))

                rowBegin += 1
                # 推送进度数据
                self.signalProgress.emit("step1ProcessWorker", [
                    "【数据处理】【第{}行】".format(rowBegin), current_percent]
                )


            # 写入折扣
            secondLoop = start_point["row"]
            for row_index_2 in range(sheet.max_row):
                if row_index_2 == 0:
                    secondLoop += 1
                    continue

                current_percent = int(((row_index_2+sheet.max_row)/(sheet.max_row * 2))  * 100)

                self.signalProgress.emit("step1ProcessWorker", ["【数据处理】【第{}行】计算渠道折扣".format(secondLoop), current_percent])

                qd_name_2_ = sheet.cell(row=secondLoop, column=QuDao).value
                qd_discount = 1.0
                # 渠道销售总额
                qdtv2_ = QuDaoTotal[qd_name_2_]

                if qd_name_2_ in configs["discount"]:
                    for rge_ in configs["discount"][qd_name_2_]:
                        if float(qdtv2_).is_integer():
                            if int(qdtv2_) in rge_[0]:
                                qd_discount = rge_[1]
                                break
                        else:
                            # range 对象的最后一个数字和销售额比对
                            if rge_[0][-1] - int(qdtv2_) > 0:
                                qd_discount = rge_[1]
                                break
                else:
                    print("渠道名称没有在配置中：{}".format(qd_name_2_))
                sheet.cell(row=secondLoop, column=QuDaoDiscount, value=qd_discount)

                self.signalProgress.emit("step1ProcessWorker", 
                    ["【数据处理】【第{}行】计算货款合计".format(secondLoop),current_percent])

                # 计算总额
                oitfv2_ = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value
                oev2_ = sheet.cell(row=rowBegin, column=OrderExtra).value
                op2_ = sheet.cell(row=rowBegin, column=OrderPostage).value
                sheet.cell(row=secondLoop, column=OrderTotal, value="=SUM(PRODUCT(K{0}, M{0}), L{0}, O{0})".format(row_index_2+ 1))
                secondLoop += 1

            # 删除列,删除后，记录的要删除的行要-1
            # init = 0
            # for del_row_num in delete_rows:
            #     sheet.delete_rows(del_row_num - init)
            #     init += 1
            #     self.signalProgress.emit("step1ProcessWorker",[
            #         "【数据处理】【第{}行】删除`订单状态`中取消的订单，`追加备注` 中带有（取消/二次配送/不接算）的订单".format(del_row_num), 
            #         current_percent]
            #     )

            result_filename = "result.xlsx"
            abs_filename = os.path.join(os.path.dirname(os.path.dirname(__file__)), result_filename)
            try:
                sfile.save(filename=abs_filename)
            except Exception as e:
                self.signalError.emit(exceptError(traceback.format_exc(), reason="保存文件失败（{}）".format(abs_filename)))
                return
            self.signalProgress.emit("step1ProcessWorker", ["【处理完成】结果暂存到：{}".format(abs_filename), 100, abs_filename])
        except Exception as e:
            self.signalError.emit(exceptError(traceback.format_exc(), reason="处理数据Excel发生了未知错误"))