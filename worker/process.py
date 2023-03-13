import traceback
import datetime
from decimal import Decimal
import re

from openpyxl import load_workbook
from PyQt5.QtCore import QThread, pyqtSignal

from error import simpleError, exceptError, baseError

class step1ProcessWorker(QThread):
    # 0: 阶段 1:进度 
    signalProgress = pyqtSignal(str, int)
    # 0: baseError
    signalError = pyqtSignal(baseError)

    def runParams(self, **params):
        self.params = params
        # params = {
        #     "source_file": self.step1SourceFileChooseRV.text(),
        #     "source_sheet": self.step1SourceDataSheetSelect.currentData(),
        #     "guessSheetXY": self.guessSheetXY,
        #     "configs": configs
        # }
    
    def run(self):
        try:
            sfile = load_workbook(filename=self.params["source_file"])
            sheet = sfile[self.params["source_sheet"]]
        except:
            self.signalError.emit(exceptError(traceback.format_exc(), reason="源文件打开失败"))
            return

        # 计算表格的起始位置
        start_point = self.params["guessSheetXY"](sheet)
        # 折扣、配件价格配置信息
        configs = self.params["configs"]

        # 首先插入要添加的列
        for title in reversed(["货款合计Fixed", "配件费用", "折扣", "结算金额"]):
            sheet.insert_cols(11)
            sheet.cell(row=start_point["row"], column=11, value=title)

        # 确定需要被处理信息坐在的列
        OrderStatus = OrderComment = 0
        OrderCreateTime = OrderDeliveryTime = 0
        OrderDetail = OrderItemTotal = OrderItemTotalFixed = OrderTotal = OrderExtra =  0
        QuDao = QuDaoDiscount =  0

        column_start = start_point["column"]
        while True:
            title_name = sheet.cell(row=start_point["row"], column=column_start).value
            if title_name =="订单状态": OrderStatus = column_start
            if title_name =="追加备注": OrderComment = column_start

            if title_name =="下单时间": OrderCreateTime = column_start
            if title_name =="承诺发货时间": OrderDeliveryTime = column_start

            if title_name =="货品摘要": OrderDetail = column_start
            if title_name =="货款合计": OrderItemTotal = column_start
            if title_name =="货款合计Fixed": OrderItemTotalFixed = column_start
            if title_name =="结算金额": OrderTotal = column_start
            if title_name =="配件费用": OrderExtra = column_start
            if title_name =="应收邮资": OrderPostage = column_start

            if title_name =="销售渠道": QuDao = column_start
            if title_name =="折扣": QuDaoDiscount = column_start
            column_start += 1
            if not title_name:
                break

        rowBegin = start_point["row"]

        # 配件价格
        peijian_restr = "|".join(configs["peijian"].keys())
        QuDaoTotal = {}
        for row_number in range(sheet.max_row):
            if row_number == 0:
                rowBegin += 1
                continue

            # 需求1：删除【订单状态】中取消的订单，【追加备注】中带有（取消/二次配送/不接算）的订单
            osv_ = sheet.cell(row=rowBegin, column=OrderStatus).value
            ocv_ = sheet.cell(row=rowBegin, column=OrderComment).value
            if osv_ :
                if "取消" in osv_:
                    sheet.delete_rows(rowBegin)
            if ocv_:
                if ("取消" in ocv_ or "不结算" in  ocv_ or "二次配送" in ocv_):
                    sheet.delete_rows(rowBegin)

            # 需求2：【下单时间】【承诺发货时间】拆分出日期，删掉时间部分
            octv__ = sheet.cell(row=rowBegin, column=OrderCreateTime).value
            odtv__ = sheet.cell(row=rowBegin, column=OrderDeliveryTime).value

            if octv__:
                dt1 = datetime.datetime.strptime(octv__, "%Y-%m-%d %H:%M:%S")
                sheet.cell(row=rowBegin, column=OrderCreateTime, value=dt1.strftime("%Y/%m/%d"))

            if odtv__:
                dt2 = datetime.datetime.strptime(odtv__, "%Y-%m-%d %H:%M:%S")
                sheet.cell(row=rowBegin, column=OrderDeliveryTime, value=dt2.strftime("%Y/%m/%d"))

            # 需求3：拆分货品摘要，计算货款合计、配件费用
            odv___ = sheet.cell(row=rowBegin, column=OrderDetail).value
            oitv___ = sheet.cell(row=rowBegin, column=OrderItemTotal).value

            # 初始化值（货款合计Fixed）
            if oitv___:
                sheet.cell(row=rowBegin, column=OrderItemTotalFixed, value=Decimal(oitv___).quantize(Decimal("0.00")))

            if odv___:
                for item in odv___.split(","):
                    if re.findall(peijian_restr, item):
                        # 匹配货品和数量
                        mapped = re.search("(.*)\[.*\]\((.*)\)", item)
                        if mapped:
                            mapped_group = mapped.groups()
                            try:
                                price = configs["peijian"][mapped_group[0]]["JiaGe"]
                                number = int(mapped_group[1])
                                # 配件费用修改
                                # 现有的值
                                oev___ = sheet.cell(row=rowBegin, column=OrderExtra).value
                                if not oev___:
                                    oev_cur = Decimal(0).quantize(Decimal("0.00"))
                                else:
                                    oev_cur = Decimal(oev___).quantize(Decimal("0.00"))
                                sheet.cell(row=rowBegin, column=OrderExtra, value=oev_cur + price * number)

                                # 修改货款合计（保存到货款合计Fixed）
                                oitfv_cur = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value
                                sheet.cell(row=rowBegin, column=OrderItemTotalFixed, value=oitfv_cur - price * number)
                            except Exception as e:
                                print(str(e))
                                # 错误处理
                        else:
                            print("格式有误！！！！！")
                            # 错误处理
                            pass

            # 需求4：计算渠道折扣
            # QuDaoDiscount
            qdv____ = sheet.cell(row=rowBegin, column=QuDao).value
            oitfv____ = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value

            if qdv____  not in QuDaoTotal:
                QuDaoTotal[qdv____] = Decimal(0).quantize(Decimal("0.00"))
            else:
                if oitfv____:
                    QuDaoTotal[qdv____] += Decimal(oitfv____).quantize(Decimal("0.00"))

            rowBegin += 1
        
        # 写入折扣
        secondLoop = start_point["row"]
        for row_number in range(sheet.max_row):
            if row_number == 0:
                secondLoop += 1
                continue

            qd_name_2_ = sheet.cell(row=secondLoop, column=QuDao).value
            qd_discount = 1.0
            # 渠道销售总额
            qdtv2_ = QuDaoTotal[qd_name_2_]

            if qd_name_2_ in configs["discount"]:
                for rge_ in configs["discount"][qd_name_2_]:
                    if float(qdtv2_).is_integer():
                        if int(qdtv2_) in rge_[0]:
                            qd_discount = rge_[1]
                            break
                    else:
                        # range 对象的最后一个数字和销售额比对
                        if rge_[0][-1] - int(qdtv2_) > 0:
                            qd_discount = rge_[1]
                            break
            sheet.cell(row=secondLoop, column=QuDaoDiscount, value=qd_discount)

            # 计算总额
            oitfv2_ = sheet.cell(row=rowBegin, column=OrderItemTotalFixed).value
            oev2_ = sheet.cell(row=rowBegin, column=OrderExtra).value
            op2_ = sheet.cell(row=rowBegin, column=OrderPostage).value
            sheet.cell(row=secondLoop, column=OrderTotal, value="=SUM(PRODUCT(K{0}, M{0}), L{0}, O{0})".format(row_number+ 1))
            secondLoop += 1
        sfile.save("new_file.xlsx")

        print("done!!!!!!!!!")
        return