import traceback
import datetime
from decimal import Decimal
import re
import os

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from PyQt5.QtCore import QThread, pyqtSignal

from error import simpleError, contentError, exceptError, baseError

class step2ProcessWorker(QThread):
    # str: workername, list: [0: 阶段, 1:进度, 2: other]
    # progres = ["stage-desc", "stage", "other"]
    signalProgress = pyqtSignal(str, list)
    # 0: baseError
    signalError = pyqtSignal(baseError)
    signalConsole = pyqtSignal(str)

    def runParams(self, **params):
        # params = {
        #     "source_configs": self.source_configs,
        #     "third_configs": self.third_configs,
        #     "guessSheetXY": self.guessSheetXY
        # }
        # source_configs = {
        #     "file_name": fname,
        #     "work_sheet": work_sheet,
        #     "order_column": 1,
        #     "amount_column": 1,
        # }
        # third_configs = {}.fromkeys(["file_name", "work_sheet", "order_column", "amount_column"])

        self.source_configs = params["source_configs"]
        self.third_configs = params["third_configs"]
        self.guessSheetXY = params["guessSheetXY"]

        # 分析数据
        self.sourceData = {}
        self.thirdData  = {}


    def run(self):
        try:
            wb = Workbook()
            
            sourceData = self.dataPreProcess(self.source_configs)
            for third_ in self.third_configs:
                thirdData = self.dataPreProcess(third_)
                if sourceData and thirdData:
                    self.dataCompare(sourceData, thirdData)
                
                if third_ == self.third_configs[-1]:
                    self.writeToExcel(wb, third_, last=True)
                else:
                    self.writeToExcel(wb, third_, last=False)
        except Exception as e:
            self.signalError.emit(exceptError(traceback.format_exc(), reason="处理数据Excel发生了未知错误"))

    def dataCompare(self, s, o):
        right_s = {}
        right_o = {}
        diff_s = {}
        diff_o = {}
        error_s = {}
        error_o = {}
        
        # 分析差异，公司订单为依据
        for k,v in s.items():
            total = Decimal(sum(v)).quantize(Decimal("0.000"))
            if k not in o:
                if k not in diff_s:
                    diff_s[k] = []
                    diff_s[k].extend(v)
                else:
                    diff_s[k].extend(v)
                self.signalConsole.emit(f"订单:{k}, 总价:{total}, 不再他方数据表中")
            else:
                anti_total = Decimal(sum(o[k])).quantize(Decimal("0.000"))
                if total != anti_total:
                    self.signalConsole.emit(f"订单:{k}, 不一致, 我方总价{total}, 他方{anti_total}")
                    if k not in error_s:
                        error_s[k] = []
                        error_s[k].extend(v)
                    else:
                        error_s[k].extend(v)
                else:
                    if k not in right_s:
                        right_s[k] = []
                        right_s[k].extend(v)
                    else:
                        right_s[k].extend(v)

        # 分析差异，三方订单为依据
        for k,v in o.items():
            total_ = Decimal(sum(v)).quantize(Decimal("0.000"))
            if k not in s:
                if k not in diff_o:
                    diff_o[k] = []
                    diff_o[k].extend(v)
                else:
                    diff_o[k].extend(v)
                self.signalConsole.emit(f"订单:{k}, 总价:{total_}, 不再我方数据表中")
            else:
                anti_total_ = Decimal(sum(s[k])).quantize(Decimal("0.000"))
                if total_ != anti_total_:
                    self.signalConsole.emit(f"订单:{k}, 不一致, 他方总价{total_}, 我方{anti_total_}")
                    if k not in error_o:
                        error_o[k] = []
                        error_o[k].extend(v)
                    else:
                        error_o[k].extend(v)
                else:
                    if k not in right_o:
                        right_o[k] = []
                        right_o[k].extend(v)
                    else:
                        right_o[k].extend(v)
        # 排序结果
        self.sort_right_s = dict(sorted(right_s.items(), key=lambda item: item[0]))
        self.sort_right_o = dict(sorted(right_o.items(), key=lambda item: item[0]))
        self.sort_diff_s = dict(sorted(diff_s.items(), key=lambda item: item[0]))
        self.sort_diff_o = dict(sorted(diff_o.items(), key=lambda item: item[0]))
        self.sort_error_s = dict(sorted(error_s.items(), key=lambda item: item[0]))
        self.sort_error_o = dict(sorted(error_o.items(), key=lambda item: item[0]))


        self.error_diff = []
        self.error_diff.extend(list(set(self.sort_error_s.keys()).difference(set(self.sort_error_o.keys()))))
        self.error_diff.extend(list(set(self.sort_error_o.keys()).difference(set(self.sort_error_s.keys()))))

        if len(self.sort_error_o) != len(self.sort_error_s) and self.error_diff:
            self.signalConsole.emit("\n\n\n")
            self.signalConsole.emit('###'*20)
            self.signalConsole.emit('不一致的订单双方数量不一致，请注意!!!')
            self.signalConsole.emit('###'*20)

    def dataPreProcess(self, config):
        dataContainer = {}
        try:
            sfile = load_workbook(filename=config["file_name"], data_only=True)
            sheet = sfile[config["work_sheet"]]
            self.signalProgress.emit("analysisWorker",
                ["【源文件处理】【{}】打开Excel文件".format(config["file_name"]), 1])
        except:
            self.signalError.emit(exceptError(traceback.format_exc(), reason="源文件打开失败"))
            return dataContainer
        else:


            start_point = self.guessSheetXY(sheet)
            rowBegin = start_point["row"]
            order_column= config["order_column"]
            amount_column= config["amount_column"]
            # 开始的行-1 代表前面有几个空行
            for row_index in range(sheet.max_row-(start_point["row"]-1)):
                if row_index == 0:
                    rowBegin += 1
                    continue

                order_value = sheet.cell(row=rowBegin, column=order_column).value
                amount_value = sheet.cell(row=rowBegin, column=amount_column).value

                if order_value and amount_value:
                    decimal_value = Decimal(str(amount_value)).quantize(Decimal("0.000"))
                    order_value_rstrip = str(order_value).rstrip()
                    if order_value_rstrip in dataContainer:
                        dataContainer[order_value_rstrip].append(float(decimal_value))
                    else:
                        dataContainer[order_value_rstrip] = []
                        dataContainer[order_value_rstrip].append(float(decimal_value))
                else:
                    self.signalConsole.emit("文件：{}\n".format(config["file_name"]))
                    self.signalConsole.emit("第{}行数据异，订单：（{}） 金额：（{}）".format(rowBegin+1, order_value, amount_value))
                rowBegin += 1

        return dataContainer

    def writeToExcel(self, wb, configs, last=False):
        result_filename = "analysis_result.xlsx"
        abs_filename = os.path.join(os.path.dirname(os.path.dirname(__file__)), result_filename)
        work_sheet = wb.create_sheet(configs["third_name"])
        row_end_1 = 1
        row_end_2 = 1
        # style_error = ('pattern: pattern solid, fore_colour  yellow')
        # style_diff = ('pattern: pattern solid, fore_colour  orange')
        style_error = PatternFill("solid", fgColor="00FFCC00")
        style_diff = PatternFill("solid", fgColor="0033CCCC")

        for k,v in self.sort_error_s.items():
            for v_ in v:
                # work_sheet.write(row_end_1, 0, k,  style_error)
                # work_sheet.write(row_end_1, 1, v_, style_error)
                work_sheet.cell(row=row_end_1, column=1, value=k).fill = style_error
                work_sheet.cell(row=row_end_1, column=2, value=v_).fill = style_error
                row_end_1 += 1

            if k in self.sort_error_o:
                for v__ in self.sort_error_o[k]:
                    # ws.write(row_end_2, 2, k, style_error)
                    # ws.write(row_end_2, 3, v__, style_error)
                    work_sheet.cell(row=row_end_2, column=3, value=k).fill = style_error
                    work_sheet.cell(row=row_end_2, column=4, value=v__).fill = style_error
                    row_end_2 += 1

            row_end_1 = row_end_2 = max(row_end_1, row_end_2)

        for err_k in self.error_diff:
            if err_k in self.sort_error_s:
                for edv in self.sort_error_s[err_k]:
                    # ws.write(row_end_1, 0, err_k,  style_error)
                    # ws.write(row_end_1, 1, edv, style_error)
                    work_sheet.cell(row=row_end_1, column=1, value=err_k).fill = style_error
                    work_sheet.cell(row=row_end_1, column=2, value=edv).fill = style_error

                    row_end_1 += 1
                    row_end_2 += 1

            if err_k in self.sort_error_o:
                for edv_ in self.sort_error_o[err_k]:
                    # ws.write(row_end_2, 2, err_k,  style_error)
                    # ws.write(row_end_2, 3, edv_, style_error)
                    work_sheet.cell(row=row_end_2, column=3, value=err_k).fill = style_error
                    work_sheet.cell(row=row_end_2, column=4, value=edv_).fill = style_error
                    row_end_2 += 1

            row_end_1 = row_end_2 = max(row_end_1, row_end_2)

        for k,v in self.sort_diff_s.items():
            for v_ in v:
                # ws.write(row_end_1, 0, k, style_diff)
                # ws.write(row_end_1, 1, v_, style_diff)
                work_sheet.cell(row=row_end_1, column=1, value=k).fill = style_diff
                work_sheet.cell(row=row_end_1, column=2, value=v_).fill = style_diff
                row_end_1 += 1


            row_end_1 = row_end_2 = max(row_end_1, row_end_2)

        for k,v in self.sort_diff_o.items():
            for v_ in v:
                # ws.write(row_end_1, 2, k, style_diff)
                # ws.write(row_end_1, 3, v_, style_diff)
                work_sheet.cell(row=row_end_1, column=3, value=k).fill = style_diff
                work_sheet.cell(row=row_end_1, column=4, value=v_).fill = style_diff
                row_end_1 += 1


            row_end_1 = row_end_2 = max(row_end_1, row_end_2)

        for k,v in self.sort_right_s.items():
            for v_ in v:
                # ws.write(row_end_1, 0, k)
                # ws.write(row_end_1, 1, v_)
                work_sheet.cell(row=row_end_1, column=1, value=k)
                work_sheet.cell(row=row_end_1, column=2, value=v_)
                row_end_1 += 1

            if k in self.sort_right_o:
                for v__ in self.sort_right_o[k]:
                    # ws.write(row_end_2, 2, k)
                    # ws.write(row_end_2, 3, v__)
                    work_sheet.cell(row=row_end_2, column=3, value=k)
                    work_sheet.cell(row=row_end_2, column=4, value=v__)
                    row_end_2 += 1

            row_end_1 = row_end_2 = max(row_end_1, row_end_2)

        if last:
            try:
                wb.save(filename=abs_filename)
            except Exception as e:
                self.signalError.emit(exceptError(traceback.format_exc(), reason="保存文件失败（{}）".format(abs_filename)))
                return
            self.signalProgress.emit("analysisWorker", ["【处理完成】结果暂存到：{}".format(abs_filename), 100, abs_filename])
